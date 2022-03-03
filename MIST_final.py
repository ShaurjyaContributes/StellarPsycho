#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar  3 11:18:16 2022

@author: shaurjyamandal
"""
from tkinter import *
import time 
import openpyxl
import random
loc = ("MIST_easy.xlsx")
wb = openpyxl.load_workbook(loc)
sheet = wb.active

window= Tk()
window.geometry("600x400")
window.title('MIST test')
window.iconbitmap('/Users/shaurjyamandal/Downloads')  
global p
timer=0
t=0
#frame = Frame(window)
global label
label = Label(window, text="Welcome to the Mist test", font=("Helvetica",18))
label.pack(side="top", fill="x", pady=100)

def next():
    cell=sheet.cell(row=random.randint(1,70),column=1).value
    label.config(text=cell)
    window.after(500,next)
    
def clicked(value):
    global p
    p=value
    
var0=IntVar()
#clicked(var0.get())  
c0 = Radiobutton(window,text="0",variable=var0, value=0, command=lambda: clicked(var0.get()))
c0.place(x=600,y=200)
c1 = Radiobutton(window,text="1",variable=var0, value=1, command=lambda: clicked(var0.get()))
c1.place(x=600,y=250)
c2 = Radiobutton(window,text="2",variable=var0, value=2, command=lambda: clicked(var0.get()))
c2.place(x=600,y=300)
c3 = Radiobutton(window,text="3",variable=var0, value=3, command=lambda: clicked(var0.get()))
c3.place(x=600,y=350)
c4 = Radiobutton(window,text="4",variable=var0, value=4, command=lambda: clicked(var0.get()))
c4.place(x=600,y=400)
c5 = Radiobutton(window,text="5",variable=var0, value=5, command=lambda: clicked(var0.get()))
c5.place(x=600,y=450)
c6 = Radiobutton(window,text="6",variable=var0, value=6, command=lambda: clicked(var0.get()))
c6.place(x=600,y=500)
c7 = Radiobutton(window,text="7",variable=var0, value=7, command=lambda: clicked(var0.get()))
c7.place(x=600,y=550)
c8 = Radiobutton(window,text="8",variable=var0, value=8, command=lambda: clicked(var0.get()))
c8.place(x=600,y=600)
c9 = Radiobutton(window,text="9",variable=var0, value=9, command=lambda: clicked(var0.get()))
c9.place(x=600,y=650)

next()
window.mainloop()
'''
i=0
while(timer<30000):
    cell=sheet.cell(row=random.randint(1,70),column=1).value
    label.config(text=cell)
    time.sleep(0.5)
    timer += 500
'''

    
    
    
    